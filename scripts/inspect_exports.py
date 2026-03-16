"""
Inspect two export xlsx files: list sheets, columns, sample rows, and summarize labor/cost/categories.
"""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

DOWNLOADS = Path(r"C:\Users\mario\Downloads")
FILE1 = DOWNLOADS / "_export_1772304267679.xlsx"
FILE2 = DOWNLOADS / "_export_1772304296699.xlsx"
OUT_DIR = Path(__file__).resolve().parent

def inspect_book(path: Path, label: str) -> dict:
    if not path.exists():
        return {"error": f"File not found: {path}"}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    info = {"path": str(path), "label": label, "sheets": []}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
            # get max row/col from dimensions or iterate
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            info["sheets"].append({"name": sheet_name, "rows": 0, "headers": [], "sample": []})
            continue
        headers = [str(c)[:80] if c is not None else "" for c in rows[0]]
        ncols = len(headers)
        sample = list(rows[1:11])
        if len(rows) > 50:
            sample += list(rows[len(rows)//2 : len(rows)//2 + 2]) + list(rows[-3:])
        info["sheets"].append({
            "name": sheet_name,
            "rows": len(rows),
            "headers": headers,
            "sample": sample,
        })
    wb.close()
    return info

def main():
    out_lines = []
    for path, label in [(FILE1, "Export 1"), (FILE2, "Export 2")]:
        out_lines.append("=" * 80)
        out_lines.append(f"FILE: {path.name}  ({label})")
        out_lines.append("=" * 80)
        info = inspect_book(path, label)
        if "error" in info:
            out_lines.append(info["error"])
            continue
        for sh in info["sheets"]:
            out_lines.append("")
            out_lines.append(f"  Sheet: {sh['name']}  (rows: {sh['rows']})")
            out_lines.append("  Headers: " + " | ".join(sh["headers"]))
            out_lines.append("  Sample rows (first 10):")
            for i, row in enumerate(sh["sample"], 1):
                out_lines.append(f"    {i}: {row}")
        out_lines.append("")
    # Extract category hierarchy from Export 1 for labor/cost sense
    out_lines.append("")
    out_lines.append("=" * 80)
    out_lines.append("CATEGORY / LABOR STRUCTURE (from Export 1)")
    out_lines.append("=" * 80)
    if FILE1.exists():
        wb1 = openpyxl.load_workbook(FILE1, read_only=True, data_only=True)
        ws1 = wb1["data"]
        rows1 = list(ws1.iter_rows(values_only=True))
        wb1.close()
        if rows1:
            headers1 = rows1[0]
            # Assume cols: Aðalflokkur, Flokkur, Liður, Aðstæður, Tók gildi, Eining (and maybe more)
            main_cats = {}
            for row in rows1[1:]:
                if not row or row[0] is None:
                    continue
                main = str(row[0]).strip()
                sub = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                item = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                if main not in main_cats:
                    main_cats[main] = {}
                if sub and sub not in main_cats[main]:
                    main_cats[main][sub] = []
                if sub and item and item not in main_cats[main][sub]:
                    main_cats[main][sub].append(item)
            for main in sorted(main_cats.keys()):
                out_lines.append(f"\n  {main}")
                for sub in sorted(main_cats[main].keys()):
                    items = main_cats[main][sub]
                    out_lines.append(f"    -> {sub}  ({len(items)} items)")
                    for it in items[:5]:
                        out_lines.append(f"        - {it}")
                    if len(items) > 5:
                        out_lines.append(f"        ... and {len(items)-5} more")
    summary_path = OUT_DIR / "export_inspection_summary.txt"
    summary_path.write_text("\n".join(out_lines), encoding="utf-8")
    print(f"Summary written to: {summary_path}")

if __name__ == "__main__":
    main()
