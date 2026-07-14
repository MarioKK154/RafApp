import openpyxl
import os

AR_DIR = r"C:\Users\mario\Desktop\AR"

def inspect():
    files = [
        "Ákvæðisgrundvöllur - Liðir.xlsx",
        "Ákvæðisgrundvöllur Allar einingar.xlsx",
        "Ákvæðisgrundvöllur Aðalflokkar.xlsx",
        "Ákvæðisgrundvöllur Álagshlutföll.xlsx"
    ]
    
    for f in files:
        path = os.path.join(AR_DIR, f)
        if not os.path.exists(path):
            print(f"File not found: {path}")
            continue
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            print(f"\nFile: {f}")
            print(f"Sheets: {wb.sheetnames}")
            for sheet in wb.sheetnames[:1]: # Check first sheet
                ws = wb[sheet]
                rows = list(ws.iter_rows(max_row=5, values_only=True))
                print(f"Sheet: {sheet}")
                for idx, r in enumerate(rows):
                    print(f"  Row {idx+1}: {r}")
            wb.close()
        except Exception as e:
            print(f"Error reading {f}: {e}")

if __name__ == "__main__":
    inspect()
