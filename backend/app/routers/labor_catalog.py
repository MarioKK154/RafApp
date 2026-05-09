# backend/app/routers/labor_catalog.py
from fastapi import APIRouter, Depends, HTTPException, status, Request, UploadFile, File
from pydantic import BaseModel
from sqlalchemy.orm import Session
from typing import Annotated, List, Optional

from .. import crud, models, schemas, security
from ..database import get_db
from ..limiter import limiter

router = APIRouter(
    prefix="/labor-catalog",
    tags=["Labor Catalog"],
    dependencies=[Depends(security.get_current_active_user)] # Basic auth required
)

DbDependency = Annotated[Session, Depends(get_db)]
CurrentUserDependency = Annotated[models.User, Depends(security.get_current_active_user)]
# Only Admins/PMs can modify the catalog (create/edit/delete items, set tenant price)
ManagerOrAdminDependency = Annotated[models.User, Depends(security.require_role(["admin", "project manager"]))]
SuperUserDependency = Annotated[models.User, Depends(security.require_superuser)]

# Helper to get item and check tenant
def get_item_for_user(item_id: int, db: DbDependency, current_user: CurrentUserDependency) -> models.LaborCatalogItem:
    """
    Helper function to retrieve a labor item while enforcing tenant isolation.
    Superusers bypass the tenant check.
    """
    effective_tenant_id = None if current_user.is_superuser else current_user.tenant_id
    db_item = crud.get_labor_catalog_item(db, item_id=item_id)
    if not db_item:
        raise HTTPException(status_code=404, detail="Labor catalog item not found or access denied.")
    return db_item

@router.post("/", response_model=schemas.LaborCatalogItemRead, status_code=status.HTTP_201_CREATED)
@limiter.limit("100/minute")
def create_labor_item(
    request: Request,
    item_data: schemas.LaborCatalogItemCreate,
    db: DbDependency,
    current_user: SuperUserDependency
):
    """
    Creates a new labor item in the global catalog (superadmin only).
    Catalog is shared; tenants set their own prices separately.
    """
    return crud.create_labor_catalog_item(db, item_data=item_data, tenant_id=None)


@router.post("/import-ar-is")
@limiter.limit("10/minute")
async def import_ar_is_global(
    request: Request,
    file: UploadFile = File(...),
    skip_duplicates: bool = True,
    db: DbDependency = None,
    current_user: SuperUserDependency = None,
):
    """
    Import ar.is labor export into the global catalog (superadmin only).
    CSV/Excel columns: Main_category, Sub_category, Item, Conditions, Effective_date, Unit_cost.
    Updates the shared catalog for all tenants; no tenant prices are set. Tenants set their own prices separately.
    """
    content = await file.read()
    filename = (file.filename or "").lower()
    if filename.endswith(".xlsx"):
        try:
            import openpyxl
            from io import BytesIO
            wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
            # ar.is Export 1 uses a sheet named "data"; prefer it, else first sheet
            ws = wb["data"] if "data" in wb.sheetnames else (wb.active or wb.worksheets[0])
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if not rows:
                raise HTTPException(status_code=400, detail="Excel file has no data.")
            headers = ["Main_category", "Sub_category", "Item", "Conditions", "Effective_date", "Unit_cost"]
            import csv as csv_mod
            from io import StringIO
            buf = StringIO()
            w = csv_mod.writer(buf)
            w.writerow(headers)
            for row in rows[1:]:  # skip header row
                if not row or len(row) < 6:
                    continue
                # Convert each cell to string (handles datetime, float, None)
                def _cell(v):
                    if v is None:
                        return ""
                    if hasattr(v, "strftime"):
                        return v.strftime("%d.%m.%Y") if hasattr(v, "strftime") else str(v)
                    return str(v).strip() if isinstance(v, str) else str(v)
                w.writerow([
                    _cell(row[0]), _cell(row[1]), _cell(row[2]), _cell(row[3]), _cell(row[4]),
                    _cell(row[5]) if len(row) > 5 else "",
                ])
            csv_content = buf.getvalue()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")
    else:
        try:
            csv_content = content.decode("utf-8")
        except UnicodeDecodeError:
            csv_content = content.decode("utf-8-sig")
    result = crud.import_labor_catalog_from_ar_is_csv(
        db, csv_content=csv_content, tenant_id=None, skip_duplicates=skip_duplicates, global_only=True
    )
    return result


@router.get("/categories")
@limiter.limit("100/minute")
def read_labor_categories(
    request: Request,
    db: DbDependency,
    current_user: CurrentUserDependency,
    lang: Optional[str] = None,
):
    """Returns category tree for menu: main_category → sub_categories with counts."""
    return crud.get_labor_catalog_categories(db, lang=lang)


@router.post("/consolidate")
@limiter.limit("5/minute")
def consolidate_catalog(
    request: Request,
    db: DbDependency,
    current_user: SuperUserDependency,
):
    """
    Merge duplicate catalog items into one per (description, main_category, sub_category).
    Each duplicate's (conditions, units_per_hour) becomes a condition variant of the kept item.
    Run once to fix existing data. Superadmin only.
    """
    return crud.consolidate_labor_catalog(db)


class ApplyTenantBasePriceBody(BaseModel):
    price: float

@router.post("/apply-tenant-base-price")
@limiter.limit("20/minute")
def apply_tenant_base_price(
    request: Request,
    body: ApplyTenantBasePriceBody,
    db: DbDependency,
    current_user: CurrentUserDependency,
):
    """
    Set your company's price for all non-hourly catalog items (per-unit / lump) to the given value.
    Does not change hourly-rate items. Tenant only (requires tenant context).
    """
    if current_user.is_superuser or current_user.tenant_id is None:
        raise HTTPException(status_code=400, detail="Only tenants can apply a base price; superadmin has no tenant.")
    if body.price < 0:
        raise HTTPException(status_code=400, detail="Price must be non-negative.")
    return crud.apply_tenant_base_price_to_non_hourly(db, tenant_id=current_user.tenant_id, price=body.price)


@router.get("/work-load-ratios", response_model=List[schemas.WorkLoadRatioRead])
@limiter.limit("100/minute")
def read_work_load_ratios(
    request: Request,
    db: DbDependency,
    current_user: CurrentUserDependency,
    active_only: bool = True,
):
    """List work load ratios (location/condition multipliers for labor)."""
    return crud.get_work_load_ratios(db, active_only=active_only)


@router.post("/import-work-load-ratios")
@limiter.limit("10/minute")
async def import_work_load_ratios_xlsx(
    request: Request,
    file: UploadFile = File(...),
    db: DbDependency = None,
    current_user: SuperUserDependency = None,
):
    """Import ar.is work load ratios Excel (sheet 'data': Númer, Lýsing, Hlutfall, Tegund, Virkt). Superadmin only."""
    content = await file.read()
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Excel (.xlsx) required.")
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb["data"] if "data" in wb.sheetnames else (wb.active or wb.worksheets[0])
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows or len(rows) < 2:
            raise HTTPException(status_code=400, detail="Excel has no data rows.")
        data = []
        for row in rows[1:]:
            if not row or len(row) < 2:
                continue
            code = str(row[0]).strip() if row[0] is not None else ""
            desc = str(row[1]).strip() if row[1] is not None else ""
            try:
                ratio = float(row[2]) if len(row) > 2 and row[2] is not None else 0.0
            except (TypeError, ValueError):
                ratio = 0.0
            ratio_type = int(row[3]) if len(row) > 3 and row[3] is not None and isinstance(row[3], (int, float)) else None
            if ratio_type is not None and not isinstance(ratio_type, int):
                try:
                    ratio_type = int(ratio_type)
                except (TypeError, ValueError):
                    ratio_type = None
            active = True
            if len(row) > 4 and row[4] is not None:
                active = bool(row[4]) if isinstance(row[4], bool) else str(row[4]).lower() in ("true", "1", "yes", "já", "j")
            data.append({"code": code, "description": desc or code, "ratio": ratio, "ratio_type": ratio_type, "is_active": active})
        result = crud.import_work_load_ratios(db, data)
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel: {str(e)}")


@router.get("/main-categories", response_model=List[schemas.LaborMainCategoryRefRead])
@limiter.limit("100/minute")
def read_main_category_refs(
    request: Request,
    db: DbDependency,
    current_user: CurrentUserDependency,
):
    """List labor main category reference (provisional basis: ALMENNT, LAGNALEIÐIR, etc.)."""
    return crud.get_labor_main_category_refs(db)


@router.post("/import-main-categories")
@limiter.limit("10/minute")
async def import_main_categories_xlsx(
    request: Request,
    file: UploadFile = File(...),
    db: DbDependency = None,
    current_user: SuperUserDependency = None,
):
    """Import ar.is main categories Excel (sheet 'data': Númer, Lýsing). Superadmin only."""
    content = await file.read()
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Excel (.xlsx) required.")
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb["data"] if "data" in wb.sheetnames else (wb.active or wb.worksheets[0])
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows or len(rows) < 2:
            raise HTTPException(status_code=400, detail="Excel has no data rows.")
        data = []
        for row in rows[1:]:
            if not row or len(row) < 2:
                continue
            code = str(row[0]).strip() if row[0] is not None else ""
            name = str(row[1]).strip() if row[1] is not None else ""
            if code:
                data.append({"code": code, "name": name or code})
        result = crud.import_labor_main_category_refs(db, data)
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel: {str(e)}")


@router.get("/", response_model=List[schemas.LaborCatalogItemRead])
@limiter.limit("100/minute")
def read_labor_items(
    request: Request,
    db: DbDependency,
    current_user: CurrentUserDependency,
    skip: int = 0,
    limit: int = 5000,
    main_category: Optional[str] = None,
    sub_category: Optional[str] = None,
    lang: Optional[str] = None,
):
    """
    Retrieves labor catalog items, optionally filtered by main_category and sub_category.
    Limit capped at 5000. Use categories endpoint to build the menu.
    """
    effective_tenant_id = None if current_user.is_superuser else current_user.tenant_id
    if effective_tenant_id is None and not current_user.is_superuser:
        raise HTTPException(status_code=403, detail="Tenant context required for non-superusers.")
    return crud.get_labor_catalog_items(
        db, tenant_id=effective_tenant_id, skip=skip, limit=limit,
        main_category=main_category, sub_category=sub_category, lang=lang,
    )

@router.get("/{item_id}/conditions", response_model=List[schemas.LaborCatalogItemConditionRead])
@limiter.limit("100/minute")
def read_labor_item_conditions(
    request: Request,
    item_id: int,
    db: DbDependency,
    current_user: CurrentUserDependency,
    lang: Optional[str] = None,
):
    """List ar.is condition variants (Eining per condition) for this catalog item."""
    get_item_for_user(item_id, db, current_user)
    return crud.get_labor_catalog_item_conditions(db, labor_catalog_item_id=item_id, lang=lang)


@router.post("/{item_id}/import-condition-variants")
@limiter.limit("20/minute")
async def import_condition_variants(
    request: Request,
    item_id: int,
    file: UploadFile = File(...),
    db: DbDependency = None,
    current_user: SuperUserDependency = None,
):
    """
    Import ar.is detail export: Excel with columns Númer, Ástæður, Tök gildi, Fell úr gildi, Eining.
    Adds condition variants for this catalog item (one Eining value per condition).
    """
    get_item_for_user(item_id, db, current_user)
    content = await file.read()
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Excel (.xlsx) required.")
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb["data"] if "data" in wb.sheetnames else (wb.active or wb.worksheets[0])
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows or len(rows) < 2:
            raise HTTPException(status_code=400, detail="Excel has no data rows.")
        data = []
        for row in rows[1:]:
            if not row or len(row) < 2:
                continue
            code = str(row[0]).strip() if row[0] is not None else ""
            if not code:
                continue
            cond_desc = str(row[1]).strip() if row[1] is not None else code
            eff = str(row[2]).strip() if len(row) > 2 and row[2] is not None else None
            end = str(row[3]).strip() if len(row) > 3 and row[3] is not None else None
            eining = None
            if len(row) > 4 and row[4] is not None:
                try:
                    eining = float(row[4])
                except (TypeError, ValueError):
                    pass
            if eff == "":
                eff = None
            if end == "":
                end = None
            data.append({
                "code": code,
                "condition_description": cond_desc,
                "effective_date": eff,
                "end_date": end,
                "units_per_hour": eining,
            })
        result = crud.import_labor_catalog_item_conditions(db, labor_catalog_item_id=item_id, rows=data)
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel: {str(e)}")


@router.get("/{item_id}", response_model=schemas.LaborCatalogItemRead)
@limiter.limit("100/minute")
def read_single_labor_item(
    request: Request,
    item_id: int,
    db: DbDependency,
    current_user: CurrentUserDependency,
    lang: Optional[str] = None,
):
    """Retrieves a single labor catalog item by ID, with tenant_price injected."""
    db_item = get_item_for_user(item_id, db, current_user)
    tenant_id = None if current_user.is_superuser else current_user.tenant_id
    tenant_price = None
    if tenant_id is not None:
        tp = db.query(models.TenantLaborPrice).filter(
            models.TenantLaborPrice.tenant_id == tenant_id,
            models.TenantLaborPrice.labor_item_id == db_item.id,
        ).first()
        if tp:
            tenant_price = tp.price
    uph_res = None
    if db_item.units_per_hour is None:
        first_var = db.query(models.LaborCatalogItemCondition).filter(
            models.LaborCatalogItemCondition.labor_catalog_item_id == db_item.id
        ).order_by(models.LaborCatalogItemCondition.code).first()
        if first_var is not None:
            uph_res = first_var.units_per_hour
    return crud.enrich_labor_item_dict(
        db_item, tenant_price=tenant_price, units_per_hour_resolved=uph_res, lang=lang
    )

@router.put("/{item_id}", response_model=schemas.LaborCatalogItemRead)
@limiter.limit("100/minute")
def update_labor_item(
    request: Request,
    item_id: int,
    item_update: schemas.LaborCatalogItemUpdate,
    db: DbDependency,
    current_user: ManagerOrAdminDependency
):
    """Updates a labor catalog item. Tenants may only set their price and unit; superadmin may edit all fields."""
    db_item = get_item_for_user(item_id, db, current_user)
    tenant_id = None if current_user.is_superuser else current_user.tenant_id
    if not current_user.is_superuser:
        item_update = schemas.LaborCatalogItemUpdate(
            default_unit_price=item_update.default_unit_price,
            unit=item_update.unit,
        )
    updated = crud.update_labor_catalog_item(db, db_item=db_item, item_update=item_update, tenant_id=tenant_id)
    tenant_price = None
    if tenant_id is not None:
        tp = db.query(models.TenantLaborPrice).filter(
            models.TenantLaborPrice.tenant_id == tenant_id,
            models.TenantLaborPrice.labor_item_id == updated.id,
        ).first()
        if tp:
            tenant_price = tp.price
    uph_res = None
    if updated.units_per_hour is None:
        first_var = db.query(models.LaborCatalogItemCondition).filter(
            models.LaborCatalogItemCondition.labor_catalog_item_id == updated.id
        ).order_by(models.LaborCatalogItemCondition.code).first()
        if first_var is not None:
            uph_res = first_var.units_per_hour
    return crud.enrich_labor_item_dict(
        updated, tenant_price=tenant_price, units_per_hour_resolved=uph_res, lang="is"
    )

@router.delete("/{item_id}", status_code=status.HTTP_204_NO_CONTENT)
@limiter.limit("100/minute")
def delete_labor_item(
    request: Request,
    item_id: int,
    db: DbDependency,
    current_user: SuperUserDependency
):
    """Deletes a labor catalog item (superadmin only; catalog is global)."""
    db_item = get_item_for_user(item_id, db, current_user)
    crud.delete_labor_catalog_item(db, db_item=db_item)
    return None